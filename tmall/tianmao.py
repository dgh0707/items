#!/usr/bin/python3
# -*- coding: utf-8 -*-

'''
@File    : taobao.py
@software: PyCharm
@Time    : 2019/2/19 9:52

'''

import requests
import os,pymysql,datetime
from lxml import etree
import time
from urllib import request
from openpyxl import load_workbook

class Spider():
    def __init__(self):
        self.conn = pymysql.connect(user='root', passwd='2018@Amber123', db='share_hunter',
                                    host='47.93.244.121', charset='utf8')
        self.cursor = self.conn.cursor()
        # self.cursor.execute('truncate dh_sys_goods;')
        # self.conn.commit()

        # 连接本地数据库
        # self.conn = pymysql.connect(user='root', passwd='123456', db='taobao', host='localhost', charset='utf8')
        # self.cursor = self.conn.cursor()
        # self.cursor.execute('truncate dh_sys_goods;')
        # self.conn.commit()

    def get_response(self,url,header,data):

        response = requests.get(url,headers=header)
        res = response.text
        html = etree.HTML(res)
        # 商品名称
        # title =''.join(html.xpath('//div[@class="tb-detail-hd"]/h1/text()')).replace("\n", "").replace("\r", "").replace("\t", "").replace("*", "").replace("/", "")
        print(data['title'])
        image_url_list = html.xpath('//*[@id="J_UlThumb"]/li/a/img/@src')
        try:
            os.mkdir('image/' + data['url_id'])
            time.sleep(0.5)
        except:
            pass
        oss_img_url_list = []
        for image_url in image_url_list:
            img_url = 'https:' + image_url.replace("_60x60q90.jpg", "")
            try:
                request.urlretrieve(img_url, 'image/' + data['url_id'] + '/' + img_url.split('/')[-1])
            except:
                print('请求错误的URL地址：%s'% (img_url,))
            time.sleep(0.5)
            print('-----正在下载商品主图图片-------')
            oss_img_url = 'https://redbag-imgs.oss-cn-beijing.aliyuncs.com/mall/2019/0311/'+data['url_id'] + '/' + img_url.split('/')[-1]
            oss_img_url_list.append(oss_img_url)
        self.save_data(data,oss_img_url_list)

    def save_data(self,data,oss_img_url_list):
        # self.insert_sql = '''insert into dh_sys_goods (cid,title,price,source,link) value ("%s","%s","%s","%s","%s")'''
        # self.cursor.execute(self.insert_sql, (data['url_id'],data['title'],data['price'],data['source'],data['to_url'],))
        # self.conn.commit()

        now = int(time.time())
        oss_img_url = ''.join(oss_img_url_list).replace('jpghttps','jpg?x-oss-process=image/quality,q_60;https')
        tit = self.cursor.execute('select id from dh_sys_goods where cid=%s', (data['url_id']))
        if tit:
            tit = self.cursor.fetchone()
        else:
            self.cursor.execute(
                """INSERT INTO dh_sys_goods (cid,title,price,source,link,created_at,updated_at,imgsrc,remark)VALUES ("%s","%s","%s","%s","%s","%s","%s","%s","%s")""" %
                (
                    data['url_id'],
                    data['title'],
                    data['price'],
                    data['source'],
                    data['to_url'],
                    now,
                    now,
                    oss_img_url,
                    data['title'],

                )
            )
            tit = self.cursor.lastrowid

        self.conn.commit()
        print('------保存成功-------')

if __name__ == '__main__':
    s = Spider()
    wb = load_workbook('11.xlsx')
    # 获取所有的sheet表格的名字
    sheets = wb.sheetnames
    # 获取第一个表格的名称
    sheet_first = sheets[0]
    # 获取特定的worksheet
    ws = wb[sheet_first]
    # 获取表格的所有的行
    rows = ws.rows
    # 获取表格所有的列
    columns = ws.columns
    data= {'url_id':"",'title':"",'price':"",'source':"淘宝联盟",'to_url':"",'auctionUrl':""}
    for row in rows:
        data['url_id'] = [col.value for col in row][0]
        data['title'] = [col.value for col in row][1]
        data['auctionUrl'] = [col.value for col in row][3]
        data['price'] = float([col.value for col in row][5])
        data['to_url'] = [col.value for col in row][10]
        # url = https://detail.tmall.com/item.htm?id=563963408680
        auctionUrl = data['auctionUrl']
        # print(auctionUrl)
        header = {
            'user-agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36',
            'referer': 'https://pub.alimama.com/promo/search/index.htm?q=%E9%9B%B6%E9%A3%9F&userType=1&_t=1550719754088&toPage=2&perPageSize=50',
            'cookie': 't=17c09943047b22674001f4068d5fe0c7; cna=BGn+EyBR2WQCAXuLGMrFrvjP; l=bBxh4KGcvHlcl8qpBOfiNuI8aO7t1IOfhsPzw4_GbICPOz1W1o4OWZZGJTYXC3GVa6nMy3yyGx6UBYL3GyUIg; cookie2=121519ab3fc74982ab39ca1fde9fc095; v=0; _tb_token_=71f533ee93e71; alimamapwag=TW96aWxsYS81LjAgKFdpbmRvd3MgTlQgNi4xOyBXaW42NDsgeDY0KSBBcHBsZVdlYktpdC81MzcuMzYgKEtIVE1MLCBsaWtlIEdlY2tvKSBDaHJvbWUvNzIuMC4zNjI2LjEwOSBTYWZhcmkvNTM3LjM2; cookie32=24c4d0a0d61710d515841732a7a881ef; alimamapw=eUt0VxcRDhBoBgdXAgRVUVEBVgwAD1cAB1QCBQUABVAFUVdXAAhTUwc%3D; cookie31=MTQ2MDQ3MTEsTXJDYXJyb3QseXVhbmxlaUBhbWJlcnN0b25lcy5jbixUQg%3D%3D; login=V32FPkk%2Fw0dUvg%3D%3D; rurl=aHR0cHM6Ly9wdWIuYWxpbWFtYS5jb20vcHJvbW8vc2VhcmNoL2luZGV4Lmh0bT9xPSVFOSU5QiVCNiVFOSVBMyU5RiZ1c2VyVHlwZT0xJl90PTE1NTA3OTk2NzI2NTI%3D; isg=BOnpzkWdkXfABq3GhsPPdy4W-JWDHt2BYhxciYveS1APUgtk0wYguMVAFLZBSnUg',

        }
        s.get_response(auctionUrl, header,data)




