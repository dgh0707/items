#!/usr/bin/python3
# -*- coding: utf-8 -*-

'''
@File    : ali.py
@software: PyCharm
@Time    : 2019/4/15 15:24

'''
import json
import os
import time

import pymysql
from openpyxl import load_workbook
import requests,re
from lxml import etree
from urllib import request


class Spider():
    def __init__(self):
        # 连接本地数据库
        self.conn = pymysql.connect(user='root', passwd='123456', db='taobao', host='localhost', charset='utf8')
        # self.conn = pymysql.connect(user='share_hunter', passwd='2018@Amber123', db='apst_share', host='47.93.244.121',charset='utf8')
        self.cursor = self.conn.cursor()

    def get_response(self,auctionUrl, dict, herder):
        res = requests.get(auctionUrl,headers = herder).text
        html = etree.HTML(res)
        data = {'id':'9','cid':'','title':'','brand':'','attribute':'','auctionurl':'','goodsthumb':'','details':''}
        # 商品的ID
        data['cid'] = dict['url_id']
        # 商品的详情页的URL地址
        data['auctionurl'] = auctionUrl
        # 商品的名称
        data['title'] = ''.join(html.xpath('//*[@id="mod-detail-title"]/h1/text()'))
        # 商品的参数属性
        attribute = html.xpath('//*[@id="mod-detail-attributes"]/div[1]/table/tbody/tr/td/text()')
        l1 = attribute[::2]
        l2 = attribute[1::2]
        ll = []
        for i in range(len(l1)):
            ll.append(l1[i] + ':' + l2[i])
        data['attribute'] = json.dumps(ll, ensure_ascii=False)
        # 商品的品牌
        list1= json.loads(data['attribute'])
        for i in list1:
            if '品牌:' in i:
                data['brand'] = i.split(':')[-1]

        # 图片下载
        # 商品主图下载
        img_url_list = html.xpath('//*[@id="dt-tab"]/div/ul/li/div/a/img/@src')
        try:
            os.mkdir('img/' + str(dict['url_id']))
            time.sleep(0.5)
        except:
            pass
        oss_imgurl_list = []
        i = 0
        for img_url in img_url_list:
            img_url = img_url.replace('60x60.jpg','800x800.jpg')
            img_name = img_url.split('/')[-1]

            request.urlretrieve(img_url, 'img/' + str(dict['url_id']) + '/' + img_name)

            print('-----正在下载商品%s主图图片-------'% dict['url_id'])
        # 拼接OSS图片地址
            img_dict = {"id": '', "imgurl": "", "status": 0, "describe": ""}
            oss_url = 'https://redbag-imgs.oss-cn-beijing.aliyuncs.com/mall/2019/1688/img/'+ str(dict['url_id']) + '/' + img_name
            img_dict['imgurl'] = oss_url
            img_dict['id'] = i + 1
            i = img_dict['id']
            oss_imgurl_list.append(img_dict)
            data['goodsthumb'] = json.dumps(oss_imgurl_list, ensure_ascii=False)
        # print(data)
        # 商品详情图下载
        image_url_dict = ''.join(re.findall('data-tfs-url="(.*?)"',res))
        print(image_url_dict)
        image_res = requests.get(image_url_dict,headers=herder).text
        # print(image_res)
        if re.findall('SHOPTOOL_POSITION_TOP_END</div>(.*?)"}',image_res,re.M):
            r = ''.join(re.findall('SHOPTOOL_POSITION_TOP_END</div>(.*?)"}',image_res,re.M))
            image_url_list = re.findall('https://cbu01([\s\S]+?).jpg',r)
        elif re.findall('https://cbu01([\s\S]+?).jpg',image_res):
            image_url_list = re.findall('https://cbu01([\s\S]+?).jpg',image_res)
        try:
            os.mkdir('image/' + str(dict['url_id']))
            time.sleep(0.5)
        except:
            pass

        oss_imageurl_list = []
        i = 0
        print(image_url_list)
        for image_url in image_url_list:
            image_url = 'https://cbu01'+image_url+'.jpg'
            # print(image_url)
            image_id = i + 1
            i = image_id
            request.urlretrieve(image_url, 'image/' + str(dict['url_id']) + '/' + str(image_id) + '.jpg')
            time.sleep(0.5)
            print('-------正在下载详情图片%d---------' % image_id)
        # 拼接OSS详情图的地址 https://redbag-imgs.oss-cn-beijing.aliyuncs.com/mall/2019/1688/image/540862882658/5.jpg
            image_dict = {"id": '', "imageurl": "", "status": 0, "describe": ""}
            oss_url = 'https://redbag-imgs.oss-cn-beijing.aliyuncs.com/mall/2019/1688/image/'+ str(dict['url_id']) + '/' + str(image_id) + '.jpg'
            image_dict['imageurl'] = oss_url
            image_dict['id'] = str(image_id)
            oss_imageurl_list.append(image_dict)
            data['details'] = json.dumps(oss_imageurl_list, ensure_ascii=False)

        self.save_data(data)

    def save_data(self,data):
        print(data)
        tit = self.cursor.execute('select id from qw_goods where goods_number=%s', (data['cid']))
        if tit:

            tit = self.cursor.fetchone()
        else:
            self.cursor.execute(
                """INSERT INTO qw_goods (cid,goods_number,goodsname,auctionurl,brand,goods_attribute,goodsthumb,details)VALUES ("%s","%s","%s","%s","%s","%s","%s","%s")""" %
                (
                    data['id'],
                    data['cid'],
                    data['title'],
                    data['auctionurl'],
                    data['brand'],
                    pymysql.escape_string(data['attribute']),
                    pymysql.escape_string(data['goodsthumb']),
                    pymysql.escape_string(data['details']),

                )
            )
            tit = self.cursor.lastrowid

        self.conn.commit()
        print('-------------%s保存成功------------'%data['cid'])


if __name__ == '__main__':
    s = Spider()
    wb = load_workbook('111.xlsx')
    # 获取所有的sheet表格的名字
    sheets = wb.sheetnames
    # 获取第一个表格的名称
    sheet_first = sheets[0]
    # 获取特定的worksheet
    ws = wb[sheet_first]
    # 获取表格的所有的行
    rows = ws.rows
    # 获取表格所有的列
    columns = ws.columns
    dict = {'url_id': '', 'auctionUrl': ''}
    for row in rows:
        dict['url_id'] = [col.value for col in row][0]
        dict['auctionUrl'] = [col.value for col in row][1]
        auctionUrl = dict['auctionUrl'].split('?')[0]
        herder = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.109 Safari/537.36',
            'cookie': 'UM_distinctid=16a070bdcf0763-00e21dfff56908-38395d0b-144000-16a070bdcf316a; cna=BGn+EyBR2WQCAXuLGMrFrvjP; ali_ab=123.139.18.141.1554896392932.4; hng=CN%7Czh-CN%7CCNY%7C156; lid=galfen2013; unb=37947416; ad_prefer="2019/04/10 19:41:11"; CNZZDATA1261052687=1295786266-1555318569-https%253A%252F%252Fdetail.1688.com%252F%7C1555318569; cookie2=1647dfc46fd4e92c2a6a41dbaa47efc2; t=8e772c6094cbb91fa09eb57fc14d330b; _tb_token_=7eee4b4e8ae07; cookie1=BqAHZZMvSfvt1WdFfDo1Bmz4RCM9PgNQ4AdbG4unez0%3D; cookie17=UNcBdgeHY7Y%3D; sg=965; csg=5a23802c; __cn_logon__=true; __cn_logon_id__=galfen2013; ali_apache_track=c_mid=b2b-37947416|c_lid=galfen2013|c_ms=1; ali_apache_tracktmp=c_w_signed=Y; _nk_=leitai9919; last_mid=b2b-37947416; CNZZDATA1253659577=1248595310-1554894373-https%253A%252F%252Fs.1688.com%252F%7C1556006865; _is_show_loginId_change_block_=b2b-37947416_false; _show_force_unbind_div_=b2b-37947416_false; _show_sys_unbind_div_=b2b-37947416_false; _show_user_unbind_div_=b2b-37947416_false; __rn_alert__=false; alicnweb=homeIdttS%3D54923864736293627752815622587963001866%7ChomeIdttSAction%3Dtrue%7Ctouch_tb_at%3D1556008530308%7Clastlogonid%3Dgalfen2013; x5sec=7b226c61707574613b32223a2261653936643937613737313236336333353662633835316434636633623263614350366b2b2b5546454f4f5171376237684d445373774561436a4d334f5451334e4445324f7a453d227d; _csrf_token=1556009609436; l=bBNce4ruvnivKDE2BOCg5uI8aN797IRAguPRwNfXi_5QzTTCuk7OlwrZyn96Vj5R_bTB4RKU8by9-etks; isg=BOXlyIyc1VqHPzGVgoVWVIOU9KHfipnZxmAg_efKtZwh_gVwr3ethRWUiCItfrFs',

        }
        print(auctionUrl)
        s.get_response(auctionUrl, dict, herder)


